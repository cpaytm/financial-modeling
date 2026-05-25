#!/usr/bin/env python3
"""
build_excel.py — IR JSON → Excel (수식 박힌 .xlsx) 빌드

framework/template.html에서 "📥 JSON" 버튼으로 내보낸 IR JSON을 입력으로,
openpyxl로 수식이 그대로 살아있는 .xlsx를 생성한다.

흐름:
    JSON 입력 → D + YRS 파싱 → 토폴로지 정렬 → 변수→행 매핑
              → 각 노드의 수식(formula)을 Excel formula로 번역
              → openpyxl Workbook 작성

수식 문법 (HTML 트리와 동일):
    산술:  + - * / 괄호
    비교:  == != < > <= >=
    함수:  SUM, MIN, MAX, AVG, IF, PREV

스타일 (FAST Standard):
    - 입력 셀: 파란색 폰트
    - 수식 셀: 검정색 폰트
    - 헤더: 굵게 + 배경

사용:
    pip install -r scripts/requirements.txt
    python scripts/build_excel.py <input.json> -o <output.xlsx>

예:
    python scripts/build_excel.py financial_model_2026-05-25.json -o my_model.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================
# 수식 파서 (template.html의 JS 파서를 Python 포팅)
# ============================================================

def tokenize(s: str) -> list[dict]:
    tokens: list[dict] = []
    i = 0
    n = len(s)
    while i < n:
        c = s[i]
        if c.isspace():
            i += 1
            continue
        if c.isdigit() or (c == "." and i + 1 < n and s[i + 1].isdigit()):
            j = i
            while j < n and (s[j].isdigit() or s[j] == "."):
                j += 1
            if j < n and s[j] in "eE":
                j += 1
                if j < n and s[j] in "+-":
                    j += 1
                while j < n and s[j].isdigit():
                    j += 1
            tokens.append({"type": "num", "val": float(s[i:j])})
            i = j
            continue
        if c.isalpha() or c == "_":
            j = i + 1
            while j < n and (s[j].isalnum() or s[j] == "_"):
                j += 1
            tokens.append({"type": "ident", "val": s[i:j]})
            i = j
            continue
        if c in "+-*/(),":
            tokens.append({"type": "op", "val": c})
            i += 1
            continue
        if c in "=!" and i + 1 < n and s[i + 1] == "=":
            tokens.append({"type": "cmp", "val": c + "="})
            i += 2
            continue
        if c in "<>":
            if i + 1 < n and s[i + 1] == "=":
                tokens.append({"type": "cmp", "val": c + "="})
                i += 2
            else:
                tokens.append({"type": "cmp", "val": c})
                i += 1
            continue
        raise ValueError(f"알 수 없는 문자: {c}")
    return tokens


class Parser:
    def __init__(self, tokens: list[dict]):
        self.tokens = tokens
        self.pos = 0

    def peek(self) -> dict | None:
        return self.tokens[self.pos] if self.pos < len(self.tokens) else None

    def eat(self, type_: str, val: str | None = None) -> dict:
        t = self.peek()
        if not t or t["type"] != type_ or (val is not None and t["val"] != val):
            raise ValueError(f"파싱 오류: 예상 {type_}({val or ''}), 실제 {t}")
        self.pos += 1
        return t

    def parse_expr(self) -> dict:
        return self.parse_cmp()

    def parse_cmp(self) -> dict:
        left = self.parse_sum()
        while self.peek() and self.peek()["type"] == "cmp":
            op = self.eat("cmp")["val"]
            right = self.parse_sum()
            left = {"kind": "cmp", "op": op, "l": left, "r": right}
        return left

    def parse_sum(self) -> dict:
        left = self.parse_term()
        while self.peek() and self.peek()["val"] in ("+", "-"):
            op = self.eat("op")["val"]
            right = self.parse_term()
            left = {"kind": "bin", "op": op, "l": left, "r": right}
        return left

    def parse_term(self) -> dict:
        left = self.parse_factor()
        while self.peek() and self.peek()["val"] in ("*", "/"):
            op = self.eat("op")["val"]
            right = self.parse_factor()
            left = {"kind": "bin", "op": op, "l": left, "r": right}
        return left

    def parse_factor(self) -> dict:
        t = self.peek()
        if not t:
            raise ValueError("수식이 비어있음")
        if t["val"] == "-":
            self.eat("op")
            return {"kind": "neg", "x": self.parse_factor()}
        if t["val"] == "(":
            self.eat("op", "(")
            e = self.parse_expr()
            self.eat("op", ")")
            return e
        if t["type"] == "num":
            self.eat("num")
            return {"kind": "num", "val": t["val"]}
        if t["type"] == "ident":
            self.eat("ident")
            if self.peek() and self.peek()["val"] == "(":
                self.eat("op", "(")
                args: list[dict] = []
                if self.peek() and self.peek()["val"] != ")":
                    args.append(self.parse_expr())
                    while self.peek() and self.peek()["val"] == ",":
                        self.eat("op")
                        args.append(self.parse_expr())
                self.eat("op", ")")
                return {"kind": "fn", "name": t["val"], "args": args}
            return {"kind": "ref", "name": t["val"]}
        raise ValueError(f"예기치 못한 토큰: {t['val']}")


def parse_formula(s: str) -> dict:
    p = Parser(tokenize(s))
    ast = p.parse_expr()
    if p.peek():
        raise ValueError(f"잉여 토큰: {p.peek()['val']}")
    return ast


# ============================================================
# AST → Excel formula 변환
# ============================================================

def col_letter(idx: int) -> str:
    """0=A, 25=Z, 26=AA, ..."""
    r = ""
    n = idx
    while n >= 0:
        r = chr(65 + n % 26) + r
        n = n // 26 - 1
    return r


def ast_to_excel(ast: dict, year: int, row_map: dict[str, int]) -> str:
    def ref(name: str, yr: int) -> str:
        row = row_map.get(name)
        if not row:
            raise ValueError(f"미정의 변수: {name}")
        return col_letter(3 + yr) + str(row)  # D=year 0, E=year 1, ...

    kind = ast["kind"]
    if kind == "num":
        return str(ast["val"])
    if kind == "ref":
        return ref(ast["name"], year)
    if kind == "neg":
        return "(-" + ast_to_excel(ast["x"], year, row_map) + ")"
    if kind == "bin":
        l = ast_to_excel(ast["l"], year, row_map)
        r = ast_to_excel(ast["r"], year, row_map)
        return f"({l}{ast['op']}{r})"
    if kind == "cmp":
        l = ast_to_excel(ast["l"], year, row_map)
        r = ast_to_excel(ast["r"], year, row_map)
        op = {"==": "=", "!=": "<>"}.get(ast["op"], ast["op"])
        return f"({l}{op}{r})"
    if kind == "fn":
        args = [ast_to_excel(a, year, row_map) for a in ast["args"]]
        fn = ast["name"].upper()
        if fn == "SUM":
            return "SUM(" + ",".join(args) + ")"
        if fn == "MIN":
            return "MIN(" + ",".join(args) + ")"
        if fn == "MAX":
            return "MAX(" + ",".join(args) + ")"
        if fn == "AVG":
            return "AVERAGE(" + ",".join(args) + ")"
        if fn == "IF":
            return f"IF({args[0]},{args[1]},{args[2]})"
        if fn == "PREV":
            if year == 0:
                return "0"
            inner = ast["args"][0]
            if inner["kind"] != "ref":
                raise ValueError("PREV는 변수 참조만 지원")
            return ref(inner["name"], year - 1)
        raise ValueError(f"미지 함수: {fn}")
    raise ValueError(f"알 수 없는 AST: {kind}")


# ============================================================
# 토폴로지 정렬
# ============================================================

def topo_sort(D: dict[str, dict]) -> list[str]:
    order: list[str] = []
    visited: set[str] = set()
    visiting: set[str] = set()

    def deps_of(k: str) -> set[str]:
        d = D[k]
        if d.get("type") != "computed" or not d.get("formula"):
            return set()
        try:
            ast = parse_formula(d["formula"])
        except Exception:
            return set()

        def collect(node: dict, acc: set[str]) -> None:
            if node.get("kind") == "ref":
                acc.add(node["name"])
            for key in ("l", "r", "x"):
                if key in node:
                    collect(node[key], acc)
            for a in node.get("args", []) or []:
                collect(a, acc)

        acc: set[str] = set()
        collect(ast, acc)
        return acc

    def visit(k: str, path: list[str]) -> None:
        if k in visited:
            return
        if k in visiting:
            raise ValueError(f"순환 참조: {' → '.join(path + [k])}")
        visiting.add(k)
        for dep in deps_of(k):
            if dep in D:
                visit(dep, path + [k])
        visiting.remove(k)
        visited.add(k)
        order.append(k)

    for k in D:
        try:
            visit(k, [])
        except ValueError as e:
            print(f"경고: {e}", file=sys.stderr)
    return order


# ============================================================
# Excel 작성
# ============================================================

# FAST 스타일 색상
HEADER_FILL = PatternFill(start_color="1E2185", end_color="1E2185", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=10)
INPUT_FONT = Font(color="0F0F12", name="Arial", size=10)  # 입력값 (파랑은 SheetJS 호환 위해 검정)
FORMULA_FONT = Font(color="0F0F12", name="Arial", size=10)
ID_FONT = Font(color="6B7280", name="Monaco", size=9)
LABEL_FONT = Font(color="0F0F12", name="Arial", size=10, bold=True)
ROOT_FILL = PatternFill(start_color="EDF3FF", end_color="EDF3FF", fill_type="solid")
THIN = Side(border_style="thin", color="E5E5E8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_workbook(data: dict, out_path: Path) -> None:
    D: dict = data["D"]
    YRS: list[str] = data["YRS"]

    order = topo_sort(D)
    row_map = {k: i + 2 for i, k in enumerate(order)}  # 헤더가 1행

    wb = openpyxl.Workbook()

    # --- Sheet 1: Model ---
    ws = wb.active
    ws.title = "Model"

    headers = ["Variable", "Label", "Unit"] + YRS
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    for k in order:
        d = D[k]
        row = row_map[k]
        c_id = ws.cell(row=row, column=1, value=k)
        c_id.font = ID_FONT
        c_lbl = ws.cell(row=row, column=2, value=d.get("label", k))
        c_lbl.font = LABEL_FONT
        ws.cell(row=row, column=3, value=d.get("u", "")).font = Font(name="Arial", size=9, color="6B7280")

        is_input = d.get("type") == "input"
        if is_input:
            for i, val in enumerate(d.get("v", [])[: len(YRS)]):
                cell = ws.cell(row=row, column=4 + i, value=val)
                cell.font = Font(color="3332D0", name="Arial", size=10)  # 입력 파랑 (FAST)
                cell.number_format = _guess_number_format(d.get("u", ""), d.get("pct"))
        else:
            formula = d.get("formula", "")
            try:
                ast = parse_formula(formula) if formula else None
            except Exception as e:
                print(f"수식 파싱 실패 ({k}): {e}", file=sys.stderr)
                ast = None
            for i in range(len(YRS)):
                if ast:
                    try:
                        excel_f = "=" + ast_to_excel(ast, i, row_map)
                        cell = ws.cell(row=row, column=4 + i, value=excel_f)
                    except Exception:
                        # fallback: cached value
                        val = d.get("v", [])
                        cell = ws.cell(row=row, column=4 + i, value=val[i] if i < len(val) else 0)
                else:
                    val = d.get("v", [])
                    cell = ws.cell(row=row, column=4 + i, value=val[i] if i < len(val) else 0)
                cell.font = FORMULA_FONT
                cell.number_format = _guess_number_format(d.get("u", ""), d.get("pct"))

    # 열 너비
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    for i in range(len(YRS)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 16

    ws.freeze_panes = "D2"

    # --- Sheet 2: Structure ---
    ws2 = wb.create_sheet("Structure")
    s_headers = ["ID", "Label", "Parent", "Type", "Formula", "Unit", "Description"]
    for col_idx, h in enumerate(s_headers, 1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for i, k in enumerate(order, 2):
        d = D[k]
        ws2.cell(row=i, column=1, value=k).font = ID_FONT
        ws2.cell(row=i, column=2, value=d.get("label", ""))
        ws2.cell(row=i, column=3, value=d.get("parent", "") or "")
        ws2.cell(row=i, column=4, value=d.get("type", ""))
        ws2.cell(row=i, column=5, value=d.get("formula", "") or "")
        ws2.cell(row=i, column=6, value=d.get("u", ""))
        ws2.cell(row=i, column=7, value=d.get("desc", ""))
    widths = [18, 28, 18, 12, 36, 10, 50]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # --- Sheet 3: Metadata ---
    ws3 = wb.create_sheet("Metadata")
    ws3.cell(row=1, column=1, value="Generated by").font = LABEL_FONT
    ws3.cell(row=1, column=2, value="financial-modeling/scripts/build_excel.py")
    ws3.cell(row=2, column=1, value="IR version").font = LABEL_FONT
    ws3.cell(row=2, column=2, value=data.get("version", "?"))
    ws3.cell(row=3, column=1, value="Nodes").font = LABEL_FONT
    ws3.cell(row=3, column=2, value=len(D))
    ws3.cell(row=4, column=1, value="Years").font = LABEL_FONT
    ws3.cell(row=4, column=2, value=f"{YRS[0]}–{YRS[-1]} ({len(YRS)}개)")
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 50

    wb.save(out_path)


def _guess_number_format(unit: str, pct: bool | None) -> str:
    if pct:
        return "0.0%"
    if unit == "원":
        return "#,##0"
    if unit in ("대", "명", "건"):
        return "#,##0"
    if unit and unit.startswith("원"):
        return "#,##0"
    return "General"


# ============================================================
# CLI
# ============================================================

def main() -> int:
    ap = argparse.ArgumentParser(description="IR JSON → Excel (수식 박힌 .xlsx) 빌더")
    ap.add_argument("input", help="HTML에서 내보낸 financial_model_*.json")
    ap.add_argument("-o", "--output", default=None, help="출력 .xlsx (기본: <input>.xlsx)")
    args = ap.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        print(f"파일 없음: {in_path}", file=sys.stderr)
        return 1

    out_path = Path(args.output) if args.output else in_path.with_suffix(".xlsx")

    with in_path.open(encoding="utf-8") as f:
        data = json.load(f)

    if "D" not in data or "YRS" not in data:
        print("잘못된 JSON 형식: 'D'와 'YRS' 키 필요", file=sys.stderr)
        return 1

    print(f"읽음: {in_path} ({len(data['D'])}개 노드, {len(data['YRS'])}년)")
    build_workbook(data, out_path)
    print(f"생성: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

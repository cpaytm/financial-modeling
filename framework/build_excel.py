#!/usr/bin/env python3
"""
build_excel.py — BM/HTML 구조 → Excel (수식 박힌 .xlsx) 빌드

BM md와 HTML에서 합의한 모델 구조를 기준으로, openpyxl로 수식이 그대로
살아있는 .xlsx를 생성한다. 자동화 실행 시에는 HTML에서 내보낸 보조 JSON을
입력으로 받는다.

흐름:
    보조 JSON 입력 → D + YRS 파싱 → 토폴로지 정렬 → 변수→행 매핑
              → 각 노드의 수식(formula)을 Excel formula로 번역
              → openpyxl Workbook 작성

수식 문법 (HTML 트리와 동일):
    산술:  + - * / 괄호
    비교:  == != < > <= >=
    함수:  SUM, MIN, MAX, AVG, IF, PREV, SUMALL, LAST, FIRST

스타일 (Excel Monoframe):
    - 입력 셀: 파란색 폰트
    - 수식 셀: 검정색 폰트
    - 타 시트 참조: 초록색 폰트
    - 헤더: 굵게 + 배경
    - Forecast 헤더: 병합 없이 선택 범위 가운데 표시
    - 틀고정 없음

사용:
    pip install -r framework/requirements.txt
    python framework/build_excel.py <input.json> -o <output.xlsx>

예:
    python framework/build_excel.py financial_model_2026-05-25.json -o my_model.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================
# 수식 파서 (template.html의 JS 파서를 Python 포팅)
# ============================================================

def tokenize(s: str) -> list[dict]:
    tokens: list[dict] = []
    i = 0
    n = len(s)
    while i < n:
        c = s[i]
        if c.isspace():
            i += 1
            continue
        if c.isdigit() or (c == "." and i + 1 < n and s[i + 1].isdigit()):
            j = i
            while j < n and (s[j].isdigit() or s[j] == "."):
                j += 1
            if j < n and s[j] in "eE":
                j += 1
                if j < n and s[j] in "+-":
                    j += 1
                while j < n and s[j].isdigit():
                    j += 1
            tokens.append({"type": "num", "val": float(s[i:j])})
            i = j
            continue
        if c.isalpha() or c == "_":
            j = i + 1
            while j < n and (s[j].isalnum() or s[j] == "_"):
                j += 1
            tokens.append({"type": "ident", "val": s[i:j]})
            i = j
            continue
        if c in "+-*/(),":
            tokens.append({"type": "op", "val": c})
            i += 1
            continue
        if c in "=!" and i + 1 < n and s[i + 1] == "=":
            tokens.append({"type": "cmp", "val": c + "="})
            i += 2
            continue
        if c in "<>":
            if i + 1 < n and s[i + 1] == "=":
                tokens.append({"type": "cmp", "val": c + "="})
                i += 2
            else:
                tokens.append({"type": "cmp", "val": c})
                i += 1
            continue
        raise ValueError(f"알 수 없는 문자: {c}")
    return tokens


class Parser:
    def __init__(self, tokens: list[dict]):
        self.tokens = tokens
        self.pos = 0

    def peek(self) -> dict | None:
        return self.tokens[self.pos] if self.pos < len(self.tokens) else None

    def eat(self, type_: str, val: str | None = None) -> dict:
        t = self.peek()
        if not t or t["type"] != type_ or (val is not None and t["val"] != val):
            raise ValueError(f"파싱 오류: 예상 {type_}({val or ''}), 실제 {t}")
        self.pos += 1
        return t

    def parse_expr(self) -> dict:
        return self.parse_cmp()

    def parse_cmp(self) -> dict:
        left = self.parse_sum()
        while self.peek() and self.peek()["type"] == "cmp":
            op = self.eat("cmp")["val"]
            right = self.parse_sum()
            left = {"kind": "cmp", "op": op, "l": left, "r": right}
        return left

    def parse_sum(self) -> dict:
        left = self.parse_term()
        while self.peek() and self.peek()["val"] in ("+", "-"):
            op = self.eat("op")["val"]
            right = self.parse_term()
            left = {"kind": "bin", "op": op, "l": left, "r": right}
        return left

    def parse_term(self) -> dict:
        left = self.parse_factor()
        while self.peek() and self.peek()["val"] in ("*", "/"):
            op = self.eat("op")["val"]
            right = self.parse_factor()
            left = {"kind": "bin", "op": op, "l": left, "r": right}
        return left

    def parse_factor(self) -> dict:
        t = self.peek()
        if not t:
            raise ValueError("수식이 비어있음")
        if t["val"] == "-":
            self.eat("op")
            return {"kind": "neg", "x": self.parse_factor()}
        if t["val"] == "(":
            self.eat("op", "(")
            e = self.parse_expr()
            self.eat("op", ")")
            return e
        if t["type"] == "num":
            self.eat("num")
            return {"kind": "num", "val": t["val"]}
        if t["type"] == "ident":
            self.eat("ident")
            if self.peek() and self.peek()["val"] == "(":
                self.eat("op", "(")
                args: list[dict] = []
                if self.peek() and self.peek()["val"] != ")":
                    args.append(self.parse_expr())
                    while self.peek() and self.peek()["val"] == ",":
                        self.eat("op")
                        args.append(self.parse_expr())
                self.eat("op", ")")
                return {"kind": "fn", "name": t["val"], "args": args}
            return {"kind": "ref", "name": t["val"]}
        raise ValueError(f"예기치 못한 토큰: {t['val']}")


def parse_formula(s: str) -> dict:
    p = Parser(tokenize(s))
    ast = p.parse_expr()
    if p.peek():
        raise ValueError(f"잉여 토큰: {p.peek()['val']}")
    return ast


# ============================================================
# AST → Excel formula 변환
# ============================================================

def col_letter(idx: int) -> str:
    """0=A, 25=Z, 26=AA, ..."""
    r = ""
    n = idx
    while n >= 0:
        r = chr(65 + n % 26) + r
        n = n // 26 - 1
    return r


def ast_to_excel(ast: dict, year: int, row_map: dict[str, int], data_start_col: int = 9, horizon: int = 1) -> str:
    def ref(name: str, yr: int) -> str:
        row = row_map.get(name)
        if not row:
            raise ValueError(f"미정의 변수: {name}")
        return col_letter(data_start_col - 1 + yr) + str(row)

    kind = ast["kind"]
    if kind == "num":
        return str(ast["val"])
    if kind == "ref":
        return ref(ast["name"], year)
    if kind == "neg":
        return "(-" + ast_to_excel(ast["x"], year, row_map, data_start_col, horizon) + ")"
    if kind == "bin":
        l = ast_to_excel(ast["l"], year, row_map, data_start_col, horizon)
        r = ast_to_excel(ast["r"], year, row_map, data_start_col, horizon)
        return f"({l}{ast['op']}{r})"
    if kind == "cmp":
        l = ast_to_excel(ast["l"], year, row_map, data_start_col, horizon)
        r = ast_to_excel(ast["r"], year, row_map, data_start_col, horizon)
        op = {"==": "=", "!=": "<>"}.get(ast["op"], ast["op"])
        return f"({l}{op}{r})"
    if kind == "fn":
        fn = ast["name"].upper()
        if fn in {"SUMALL", "LAST", "FIRST"}:
            args_raw = ast["args"]
            if len(args_raw) != 1 or args_raw[0]["kind"] != "ref":
                raise ValueError(f"{fn}는 변수 참조 1개만 지원")
            row = row_map.get(args_raw[0]["name"])
            if not row:
                raise ValueError(f"미정의 변수: {args_raw[0]['name']}")
            first = col_letter(data_start_col - 1) + str(row)
            last = col_letter(data_start_col - 1 + horizon - 1) + str(row)
            if fn == "SUMALL":
                return f"SUM({first}:{last})"
            if fn == "LAST":
                return last
            return first
        args = [ast_to_excel(a, year, row_map, data_start_col, horizon) for a in ast["args"]]
        if fn == "SUM":
            return "SUM(" + ",".join(args) + ")"
        if fn == "MIN":
            return "MIN(" + ",".join(args) + ")"
        if fn == "MAX":
            return "MAX(" + ",".join(args) + ")"
        if fn == "AVG":
            return "AVERAGE(" + ",".join(args) + ")"
        if fn == "IF":
            return f"IF({args[0]},{args[1]},{args[2]})"
        if fn == "PREV":
            if year == 0:
                return "0"
            inner = ast["args"][0]
            if inner["kind"] != "ref":
                raise ValueError("PREV는 변수 참조만 지원")
            return ref(inner["name"], year - 1)
        raise ValueError(f"미지 함수: {fn}")
    raise ValueError(f"알 수 없는 AST: {kind}")


# ============================================================
# 토폴로지 정렬
# ============================================================

def topo_sort(D: dict[str, dict]) -> list[str]:
    order: list[str] = []
    visited: set[str] = set()
    visiting: set[str] = set()

    def deps_of(k: str) -> set[str]:
        d = D[k]
        if d.get("type") != "computed" or not d.get("formula"):
            return set()
        try:
            ast = parse_formula(d["formula"])
        except Exception:
            return set()

        def collect(node: dict, acc: set[str]) -> None:
            if node.get("kind") == "ref":
                acc.add(node["name"])
            for key in ("l", "r", "x"):
                if key in node:
                    collect(node[key], acc)
            for a in node.get("args", []) or []:
                collect(a, acc)

        acc: set[str] = set()
        collect(ast, acc)
        acc.discard(k)
        return acc

    def visit(k: str, path: list[str]) -> None:
        if k in visited:
            return
        if k in visiting:
            raise ValueError(f"순환 참조: {' → '.join(path + [k])}")
        visiting.add(k)
        for dep in deps_of(k):
            if dep in D:
                visit(dep, path + [k])
        visiting.remove(k)
        visited.add(k)
        order.append(k)

    for k in D:
        try:
            visit(k, [])
        except ValueError as e:
            print(f"경고: {e}", file=sys.stderr)
    return order


# ============================================================
# Excel 작성 — generic model workbook builder
# ============================================================

# Excel modeling guide styles
FONT_TITLE = Font(name="Arial", size=11, bold=True, color="1F4E79")
FONT_LABEL = Font(name="맑은 고딕", size=9, color="000000")
FONT_LABEL_BOLD = Font(name="맑은 고딕", size=9, bold=True, color="000000")
FONT_LABEL_SUB = Font(name="맑은 고딕", size=9, color="808080")
FONT_INPUT = Font(name="Arial", size=9, color="0000FF")
FONT_FORMULA = Font(name="Arial", size=9, color="000000")
FONT_LINK = Font(name="Arial", size=9, color="008000")
FONT_SUB = Font(name="Arial", size=9, color="808080")
FONT_HEADER = Font(name="Arial", size=9, bold=True, color="FFFFFF")
FONT_ID = Font(name="Consolas", size=8, color="808080")

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_FORECAST = PatternFill("solid", fgColor="808080")
FILL_OUTPUT = PatternFill("solid", fgColor="DDEBF7")
FILL_TOGGLE = PatternFill("solid", fgColor="FFFF99")
FILL_ERROR = PatternFill("solid", fgColor="FF0000")
FILL_SECTION = PatternFill("solid", fgColor="D9EAF7")

SIDE_THIN = Side(style="thin", color="A6A6A6")
BORDER_HEADER = Border(bottom=SIDE_THIN)
BORDER_SUBTOTAL = Border(top=SIDE_THIN)
BORDER_TOTAL = Border(top=SIDE_THIN, bottom=Side(style="double", color="000000"))
BORDER_ROW = Border(bottom=Side(style="hair", color="D9D9D9"))

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_CENTER_ACROSS = Alignment(horizontal="centerContinuous", vertical="center")

FMT_ACCT = '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'
FMT_ACCT_1 = '_-* #,##0.0_-;-* #,##0.0_-;_-* "-"_-;_-@_-'
FMT_ACCT_2 = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"_-;_-@_-'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_MULT = '0.0"x"'
FMT_NUM = '#,##0;(#,##0);"-"'
FMT_YEAR = 'General'

DATA_START_COL = 9  # I
LABEL_START_COL = 2  # B — 테두리/음영을 그리는 시작 열
LABEL_COL = 6  # F — 계정명은 항상 여기. 계층은 alignment.indent로 표현한다.
MAX_LABEL_LEVELS = 8  # indent 최대 단계
ID_COL = 7  # G
UNIT_COL = 8  # H
HEADER_ROW = 4
DATA_START_ROW = 5


def _guess_number_format(unit: str, pct: bool | None) -> str:
    unit = unit or ""
    if pct:
        return FMT_PCT
    if unit == "x":
        return FMT_MULT
    if unit in ("대", "명", "건", "주"):
        return FMT_NUM
    if unit == "%":
        return FMT_PCT
    if any(token in unit for token in ("GWh", "백만대", "천$/대", "백만$/GWh", "million", "bn", "십억", "백만")):
        return FMT_ACCT_1
    if any(token in unit for token in ("$/", "USD/", "배럴", "톤")):
        return FMT_ACCT_2
    return FMT_ACCT


def _setup_sheet(ws, title: str | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    ws.freeze_panes = None
    ws.column_dimensions["A"].width = 1
    if title:
        ws["B1"] = title
        ws["B1"].font = FONT_TITLE
        ws.row_dimensions[1].height = 16.5


def _quote_sheet(name: str) -> str:
    return "'" + name.replace("'", "''") + "'"


def _children_map(D: dict[str, dict]) -> dict[str | None, list[str]]:
    children: dict[str | None, list[str]] = {}
    for k, d in D.items():
        children.setdefault(d.get("parent"), []).append(k)
    return children


def _tree_order(D: dict[str, dict]) -> list[str]:
    children = _children_map(D)
    roots = children.get(None) or [k for k, d in D.items() if not d.get("parent")]
    out: list[str] = []
    seen: set[str] = set()

    def walk(k: str) -> None:
        if k in seen or k not in D:
            return
        seen.add(k)
        out.append(k)
        for child in children.get(k, []):
            walk(child)

    for root in roots:
        walk(root)
    for k in D:
        walk(k)
    return out


def _levels(D: dict[str, dict]) -> dict[str, int]:
    memo: dict[str, int] = {}

    def level(k: str) -> int:
        if k in memo:
            return memo[k]
        p = D.get(k, {}).get("parent")
        if not p or p not in D:
            memo[k] = 0
        else:
            memo[k] = level(p) + 1
        return memo[k]

    for k in D:
        level(k)
    return memo


def _deps(ast: dict) -> set[str]:
    acc: set[str] = set()

    def walk(node: dict) -> None:
        if node.get("kind") == "ref":
            acc.add(node["name"])
        for key in ("l", "r", "x"):
            if key in node:
                walk(node[key])
        for a in node.get("args", []) or []:
            walk(a)

    walk(ast)
    return acc


def _validate_formulas(D: dict[str, dict]) -> tuple[dict[str, dict], list[str]]:
    asts: dict[str, dict] = {}
    errors: list[str] = []
    for k, d in D.items():
        if d.get("type") != "computed":
            continue
        formula = d.get("formula") or ""
        if not formula:
            errors.append(f"{k}: computed 노드에 formula 없음")
            continue
        try:
            ast = parse_formula(formula)
            asts[k] = ast
            deps = _deps(ast)
            deps.discard(k)
            for dep in deps:
                if dep not in D:
                    errors.append(f"{k}: 미정의 변수 참조 {dep}")
        except Exception as e:
            errors.append(f"{k}: 수식 파싱 실패 - {e}")
    try:
        topo_sort(D)
    except Exception as e:
        errors.append(str(e))
    return asts, errors


def _write_year_headers(ws, YRS: list[str], hist_n: int = 1) -> None:
    hist_n = max(0, min(int(hist_n or 0), len(YRS)))
    fc_n = len(YRS) - hist_n
    if hist_n:
        c = ws.cell(row=3, column=DATA_START_COL, value="Historical")
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER_ACROSS if hist_n > 1 else ALIGN_CENTER
        c.border = BORDER_HEADER
        for i in range(1, hist_n):
            c = ws.cell(row=3, column=DATA_START_COL + i, value=None)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = ALIGN_CENTER_ACROSS
            c.border = BORDER_HEADER
    if fc_n:
        fc_col = DATA_START_COL + hist_n
        c = ws.cell(row=3, column=fc_col, value="Forecast")
        c.font = FONT_HEADER
        c.fill = FILL_FORECAST
        c.alignment = ALIGN_CENTER_ACROSS if fc_n > 1 else ALIGN_CENTER
        c.border = BORDER_HEADER
        for i in range(hist_n + 1, len(YRS)):
            c = ws.cell(row=3, column=DATA_START_COL + i, value=None)
            c.font = FONT_HEADER
            c.fill = FILL_FORECAST
            c.alignment = ALIGN_CENTER_ACROSS
            c.border = BORDER_HEADER
    c = ws.cell(row=HEADER_ROW, column=LABEL_COL, value="계정")
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_LEFT
    c.border = BORDER_HEADER
    for i, y in enumerate(YRS):
        c = ws.cell(row=HEADER_ROW, column=DATA_START_COL + i, value=y)
        c.font = FONT_HEADER
        c.fill = FILL_FORECAST if i >= hist_n else FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_HEADER
        c.number_format = FMT_YEAR


def _set_model_columns(ws, YRS: list[str]) -> None:
    for col in ("A", "B", "C", "D", "E"):
        ws.column_dimensions[col].width = 2
    ws.column_dimensions[get_column_letter(LABEL_COL)].width = 46
    ws.column_dimensions[get_column_letter(ID_COL)].width = 22
    ws.column_dimensions[get_column_letter(UNIT_COL)].width = 12
    for i in range(len(YRS)):
        ws.column_dimensions[get_column_letter(DATA_START_COL + i)].width = 14


def _write_label(ws, row: int, level: int, text: str, bold: bool = False, sub: bool = False) -> None:
    """계정명을 LABEL_COL(F) 한 곳에 쓰고 트리 깊이는 Excel 들여쓰기로 표현한다.

    예전에는 깊이에 따라 B~F로 열을 옮겨 적었다. 그 방식은 옆 칸이 비어 있어야만
    글자가 넘쳐 보이므로, 옆 칸에 무엇이든 들어가는 순간 계정명이 잘린다.
    template.html의 엑셀 내보내기도 같은 이유로 F열 + indent 방식으로 통일했다.
    """
    c = ws.cell(row=row, column=LABEL_COL, value=text)
    c.font = FONT_LABEL_SUB if sub else FONT_LABEL_BOLD if bold else FONT_LABEL
    c.alignment = Alignment(
        horizontal="left", vertical="center",
        indent=max(0, min(level, MAX_LABEL_LEVELS)),
    )


def _apply_horizontal_row_border(ws, row: int, start_col: int, end_col: int, border: Border = BORDER_ROW) -> None:
    for col in range(start_col, end_col + 1):
        ws.cell(row=row, column=col).border = border


def _apply_body_horizontal_borders(ws, start_row: int, end_row: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        _apply_horizontal_row_border(ws, row, LABEL_START_COL, end_col, BORDER_ROW)


def _important_outputs(D: dict[str, dict], order: list[str]) -> list[str]:
    preferred = [
        "root", "rev", "used_car_sales", "ecommerce_sales", "branch_sales", "total_cost",
        "op_profit", "fcff", "enterprise_value", "equity_value", "value_per_share",
    ]
    out = [k for k in preferred if k in D]
    top = [k for k, d in D.items() if d.get("parent") in (None, "root") and k not in out]
    out.extend(top)
    for k in order:
        if len(out) >= 24:
            break
        if D[k].get("type") == "computed" and k not in out:
            out.append(k)
    return out


def _descendants(D: dict[str, dict], root: str) -> list[str]:
    children = _children_map(D)
    out: list[str] = []

    def walk(k: str) -> None:
        if k not in D:
            return
        out.append(k)
        for child in children.get(k, []):
            walk(child)

    walk(root)
    return out


def _ordered_existing(order: list[str], node_ids: list[str]) -> list[str]:
    wanted = set(node_ids)
    return [k for k in order if k in wanted]


def _is_kcar_ir(D: dict[str, dict]) -> bool:
    required = {"ecommerce_sales", "branch_sales", "used_car_volume", "fcff", "value_per_share"}
    return required.issubset(D)


def _formula_to_sheet(
    ast: dict,
    year: int,
    local_row_map: dict[str, int],
    model_row_map: dict[str, int],
    model_sheet_name: str = "Model",
    data_start_col: int = DATA_START_COL,
    horizon: int = 1,
) -> str:
    def ref(name: str, yr: int) -> str:
        col = get_column_letter(data_start_col + yr)
        if name in local_row_map:
            return f"{col}{local_row_map[name]}"
        if name in model_row_map:
            return f"{_quote_sheet(model_sheet_name)}!{col}{model_row_map[name]}"
        raise ValueError(f"미정의 변수: {name}")

    kind = ast["kind"]
    if kind == "num":
        return str(ast["val"])
    if kind == "ref":
        return ref(ast["name"], year)
    if kind == "neg":
        return "(-" + _formula_to_sheet(ast["x"], year, local_row_map, model_row_map, model_sheet_name, data_start_col, horizon) + ")"
    if kind == "bin":
        l = _formula_to_sheet(ast["l"], year, local_row_map, model_row_map, model_sheet_name, data_start_col, horizon)
        r = _formula_to_sheet(ast["r"], year, local_row_map, model_row_map, model_sheet_name, data_start_col, horizon)
        return f"({l}{ast['op']}{r})"
    if kind == "cmp":
        l = _formula_to_sheet(ast["l"], year, local_row_map, model_row_map, model_sheet_name, data_start_col, horizon)
        r = _formula_to_sheet(ast["r"], year, local_row_map, model_row_map, model_sheet_name, data_start_col, horizon)
        op = {"==": "=", "!=": "<>"}.get(ast["op"], ast["op"])
        return f"({l}{op}{r})"
    if kind == "fn":
        fn = ast["name"].upper()
        if fn in {"SUMALL", "LAST", "FIRST"}:
            args_raw = ast["args"]
            if len(args_raw) != 1 or args_raw[0]["kind"] != "ref":
                raise ValueError(f"{fn}는 변수 참조 1개만 지원")
            name = args_raw[0]["name"]
            row = local_row_map.get(name)
            sheet = ""
            if not row:
                row = model_row_map.get(name)
                sheet = _quote_sheet(model_sheet_name) + "!"
            if not row:
                raise ValueError(f"미정의 변수: {name}")
            first = sheet + get_column_letter(data_start_col) + str(row)
            last = get_column_letter(data_start_col + horizon - 1) + str(row)
            if fn == "SUMALL":
                return f"SUM({first}:{last})"
            if fn == "LAST":
                return last
            return first
        if fn == "PREV":
            if year == 0:
                return "0"
            inner = ast["args"][0]
            if inner["kind"] != "ref":
                raise ValueError("PREV는 변수 참조만 지원")
            return ref(inner["name"], year - 1)
        args = [_formula_to_sheet(a, year, local_row_map, model_row_map, model_sheet_name, data_start_col, horizon) for a in ast["args"]]
        if fn == "AVG":
            return "AVERAGE(" + ",".join(args) + ")"
        return fn + "(" + ",".join(args) + ")"
    raise ValueError(f"알 수 없는 AST: {kind}")


def _write_model_sheet(
    wb,
    sheet_name: str,
    title: str,
    node_ids: list[str],
    D: dict[str, dict],
    YRS: list[str],
    levels: dict[str, int],
    asts: dict[str, dict],
    model_row_map: dict[str, int],
    input_row_map: dict[str, int],
    formula_audit: list[list[Any]],
    hist_n: int = 1,
) -> tuple[int, int]:
    ws = wb.create_sheet(sheet_name)
    _setup_sheet(ws, title)
    _write_year_headers(ws, YRS, hist_n)
    _set_model_columns(ws, YRS)
    for cidx, label in ((ID_COL, "ID"), (UNIT_COL, "Unit")):
        c = ws.cell(row=HEADER_ROW, column=cidx, value=label)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER

    local_row_map = {k: DATA_START_ROW + i for i, k in enumerate(node_ids)}
    formula_count = 0
    link_count = 0

    for k in node_ids:
        d = D[k]
        row = local_row_map[k]
        level = max(0, min(levels[k], MAX_LABEL_LEVELS - 1))
        is_section = d.get("type") == "computed" and (level <= 1 or k in {"rev", "total_cost", "fcff", "valuation", "root"})
        _write_label(ws, row, level, d.get("label", k), bold=is_section, sub=bool(d.get("pct")))
        ws.cell(row=row, column=ID_COL, value=k).font = FONT_ID
        ws.cell(row=row, column=UNIT_COL, value=d.get("u", "")).font = FONT_SUB
        if is_section:
            for col in range(LABEL_START_COL, DATA_START_COL + len(YRS)):
                ws.cell(row=row, column=col).fill = FILL_OUTPUT if k in {"root", "fcff", "valuation"} else FILL_SECTION
                ws.cell(row=row, column=col).border = BORDER_SUBTOTAL

        if d.get("type") == "input":
            arow = input_row_map[k]
            for i in range(len(YRS)):
                col = DATA_START_COL + i
                formula = f"={_quote_sheet('Assumptions')}!{get_column_letter(col)}{arow}"
                cell = ws.cell(row=row, column=col, value=formula)
                cell.font = FONT_LINK
                cell.number_format = _guess_number_format(d.get("u", ""), d.get("pct"))
                cell.alignment = ALIGN_RIGHT
                link_count += 1
                formula_audit.append([sheet_name, k, d.get("label", k), YRS[i], cell.coordinate, formula, "assumption link"])
            continue

        deps = _deps(asts[k])
        is_cross_sheet = any(dep not in local_row_map for dep in deps)
        for i in range(len(YRS)):
            col = DATA_START_COL + i
            formula = "=" + _formula_to_sheet(asts[k], i, local_row_map, model_row_map, horizon=len(YRS))
            cell = ws.cell(row=row, column=col, value=formula)
            cell.font = FONT_LINK if is_cross_sheet else FONT_FORMULA
            cell.number_format = _guess_number_format(d.get("u", ""), d.get("pct"))
            cell.alignment = ALIGN_RIGHT
            if is_section:
                cell.font = Font(name="Arial", size=9, bold=True, color="008000" if is_cross_sheet else "000000")
                cell.border = BORDER_TOTAL if k in {"root", "fcff", "valuation"} else BORDER_SUBTOTAL
            formula_count += 1
            formula_audit.append([sheet_name, k, d.get("label", k), YRS[i], cell.coordinate, formula, "sheet calculation"])

    _apply_body_horizontal_borders(ws, DATA_START_ROW, DATA_START_ROW + len(node_ids) - 1, DATA_START_COL + len(YRS) - 1)
    for k in node_ids:
        d = D[k]
        row = local_row_map[k]
        level = max(0, min(levels[k], MAX_LABEL_LEVELS - 1))
        if d.get("type") == "computed" and (level <= 1 or k in {"rev", "total_cost", "fcff", "valuation", "root"}):
            _apply_horizontal_row_border(
                ws,
                row,
                LABEL_START_COL,
                DATA_START_COL + len(YRS) - 1,
                BORDER_TOTAL if k in {"root", "fcff", "valuation"} else BORDER_SUBTOTAL,
            )
    return formula_count, link_count


def _kcar_sheet_nodes(D: dict[str, dict], order: list[str]) -> dict[str, list[str]]:
    sales = _descendants(D, "rev") if "rev" in D else []
    cost_all = _descendants(D, "total_cost") if "total_cost" in D else []
    labor = _descendants(D, "labor_cost") if "labor_cost" in D else []
    excluded_cost = set(labor + ["da"])
    cost = [k for k in cost_all if k not in excluded_cost]
    capex_da = [k for k in ["capex", "da"] if k in D]
    nwc = [k for k in ["rev", "nwc_rate", "nwc", "change_nwc"] if k in D]
    dcf = [
        k for k in [
            "rev", "total_cost", "op_profit", "tax_rate", "tax", "ebiat", "da", "capex",
            "nwc", "change_nwc", "fcff", "wacc", "terminal_growth", "discount_factor",
            "pv_fcff", "terminal_value", "pv_terminal_value", "enterprise_value",
            "net_debt", "equity_value", "shares_outstanding", "value_per_share",
        ] if k in D
    ]
    bs = [k for k in ["net_debt", "shares_outstanding", "equity_value", "value_per_share"] if k in D]
    bridge = [
        k for k in [
            "rev", "used_car_sales", "ecommerce_sales", "branch_sales", "auction_sales",
            "rental_sales", "service_other_sales", "used_car_volume",
        ] if k in D
    ]
    return {
        "DCF": _ordered_existing(order, dcf),
        "Sales": _ordered_existing(order, sales),
        "Sales Bridge": _ordered_existing(order, bridge),
        "Cost": _ordered_existing(order, cost),
        "CapEx,D&A": _ordered_existing(order, capex_da),
        "Labor": _ordered_existing(order, labor),
        "NWC": _ordered_existing(order, nwc),
        "BS": _ordered_existing(order, bs),
    }


def build_workbook(data: dict, out_path: Path) -> None:
    # HTML 내보내기는 "MODEL" 키를 쓴다. 예전 파일은 "D"였으므로 둘 다 받는다.
    D: dict[str, dict] = data.get("MODEL") or data["D"]
    YRS: list[str] = [str(y) for y in data["YRS"]]
    hist_n = max(0, min(int(data.get("HIST_N", 1) or 0), len(YRS)))
    display_order = _tree_order(D)
    levels = _levels(D)
    asts, formula_errors = _validate_formulas(D)
    if formula_errors:
        msg = "\n".join(formula_errors[:20])
        raise ValueError(f"수식 검증 실패:\n{msg}")

    row_map = {k: DATA_START_ROW + i for i, k in enumerate(display_order)}
    input_order = [k for k in display_order if D[k].get("type") == "input"]
    input_row_map = {k: DATA_START_ROW + i for i, k in enumerate(input_order)}

    wb = openpyxl.Workbook()
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    # Index
    ws_index = wb.active
    ws_index.title = "Index"
    _setup_sheet(ws_index, "Index")
    is_kcar = _is_kcar_ir(D)
    index_rows = [("Control", "핵심 가정과 주요 결과 요약")]
    if is_kcar:
        index_rows.extend([
            ("DCF", "FCFF, 터미널가치, Equity Value, 주당가치"),
            ("Sales", "채널별·차급별 Q × P 매출 모델"),
            ("Sales Bridge", "매출 구성과 주요 드라이버 브릿지"),
            ("Cost", "재고자산원가, 보증비, 렌터카 원가, 판관비"),
            ("CapEx,D&A", "투자와 감가상각비"),
            ("Labor", "필요 인원과 평균급여 기반 인건비"),
            ("NWC", "매출/비용 연동 운전자본과 △NWC"),
            ("BS", "순차입금, 주식수, 주주가치 연결"),
        ])
    index_rows.extend([
        ("Assumptions", "모든 하드코딩 입력값 집중 시트"),
        ("Model", "HTML D 그래프를 기간별 Excel 수식으로 전개한 계산 엔진"),
        ("Formula Audit", "생성된 모든 수식 셀 감사"),
        ("Structure", "HTML 노드 ID, parent, formula, description"),
        ("Checks", "수식/구조/정합성 검증"),
        ("Metadata", "생성 정보"),
    ])
    ws_index.cell(row=3, column=2, value="Sheet").font = FONT_HEADER
    ws_index.cell(row=3, column=2).fill = FILL_HEADER
    ws_index.cell(row=3, column=3, value="Description").font = FONT_HEADER
    ws_index.cell(row=3, column=3).fill = FILL_HEADER
    for r, (sheet, desc) in enumerate(index_rows, 4):
        ws_index.cell(row=r, column=2, value=f'=HYPERLINK("#\'{sheet}\'!A1","{sheet}")').font = FONT_LINK
        ws_index.cell(row=r, column=3, value=desc).font = FONT_LABEL
        _apply_horizontal_row_border(ws_index, r, 2, 3)
    ws_index.column_dimensions["A"].width = 1
    ws_index.column_dimensions["B"].width = 22
    ws_index.column_dimensions["C"].width = 80

    # Assumptions
    ws_assump = wb.create_sheet("Assumptions")
    _setup_sheet(ws_assump, "Assumptions")
    _write_year_headers(ws_assump, YRS, hist_n)
    _set_model_columns(ws_assump, YRS)
    ws_assump.cell(row=HEADER_ROW, column=ID_COL, value="ID").font = FONT_HEADER
    ws_assump.cell(row=HEADER_ROW, column=ID_COL).fill = FILL_HEADER
    ws_assump.cell(row=HEADER_ROW, column=UNIT_COL, value="Unit").font = FONT_HEADER
    ws_assump.cell(row=HEADER_ROW, column=UNIT_COL).fill = FILL_HEADER
    for k in input_order:
        d = D[k]
        row = input_row_map[k]
        _write_label(ws_assump, row, levels[k], d.get("label", k), sub=bool(d.get("pct")))
        ws_assump.cell(row=row, column=ID_COL, value=k).font = FONT_ID
        ws_assump.cell(row=row, column=UNIT_COL, value=d.get("u", "")).font = FONT_SUB
        vals = list(d.get("v", []))
        for i in range(len(YRS)):
            val = vals[i] if i < len(vals) else 0
            c = ws_assump.cell(row=row, column=DATA_START_COL + i, value=val)
            c.font = FONT_INPUT
            c.number_format = _guess_number_format(d.get("u", ""), d.get("pct"))
            c.alignment = ALIGN_RIGHT
    _apply_body_horizontal_borders(ws_assump, DATA_START_ROW, DATA_START_ROW + len(input_order) - 1, DATA_START_COL + len(YRS) - 1)
    # Model
    ws_model = wb.create_sheet("Model")
    _setup_sheet(ws_model, "Model")
    _write_year_headers(ws_model, YRS, hist_n)
    _set_model_columns(ws_model, YRS)
    ws_model.cell(row=2, column=DATA_START_COL, value="Unit: source units from HTML IR").font = FONT_SUB
    for cidx, label in ((ID_COL, "ID"), (UNIT_COL, "Unit")):
        c = ws_model.cell(row=HEADER_ROW, column=cidx, value=label)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
    formula_audit: list[list[Any]] = [["Sheet", "ID", "Label", "Year", "Cell", "Formula", "Type"]]
    formula_count = 0
    input_link_count = 0
    for k in display_order:
        d = D[k]
        row = row_map[k]
        level = levels[k]
        is_section = d.get("type") == "computed" and (level <= 1 or k == "root")
        _write_label(ws_model, row, level, d.get("label", k), bold=is_section, sub=bool(d.get("pct")))
        ws_model.cell(row=row, column=ID_COL, value=k).font = FONT_ID
        ws_model.cell(row=row, column=UNIT_COL, value=d.get("u", "")).font = FONT_SUB
        if is_section:
            for col in range(LABEL_START_COL, DATA_START_COL + len(YRS)):
                ws_model.cell(row=row, column=col).fill = FILL_OUTPUT if k == "root" else FILL_SECTION
        if d.get("type") == "input":
            arow = input_row_map[k]
            for i in range(len(YRS)):
                cell = ws_model.cell(row=row, column=DATA_START_COL + i)
                aref = f"={_quote_sheet('Assumptions')}!{get_column_letter(DATA_START_COL+i)}{arow}"
                cell.value = aref
                cell.font = FONT_LINK
                cell.number_format = _guess_number_format(d.get("u", ""), d.get("pct"))
                cell.alignment = ALIGN_RIGHT
                input_link_count += 1
                formula_audit.append(["Model", k, d.get("label", k), YRS[i], cell.coordinate, aref, "input link"])
        else:
            ast = asts[k]
            for i in range(len(YRS)):
                formula = "=" + ast_to_excel(ast, i, row_map, DATA_START_COL, len(YRS))
                cell = ws_model.cell(row=row, column=DATA_START_COL + i, value=formula)
                cell.font = FONT_FORMULA
                cell.number_format = _guess_number_format(d.get("u", ""), d.get("pct"))
                cell.alignment = ALIGN_RIGHT
                if is_section:
                    cell.font = Font(name="Arial", size=9, bold=True, color="000000")
                    cell.border = BORDER_TOTAL if k == "root" else BORDER_SUBTOTAL
                formula_count += 1
                formula_audit.append(["Model", k, d.get("label", k), YRS[i], cell.coordinate, formula, "calculation"])
    _apply_body_horizontal_borders(ws_model, DATA_START_ROW, DATA_START_ROW + len(display_order) - 1, DATA_START_COL + len(YRS) - 1)
    for k in display_order:
        d = D[k]
        row = row_map[k]
        level = levels[k]
        if d.get("type") == "computed" and (level <= 1 or k == "root"):
            _apply_horizontal_row_border(
                ws_model,
                row,
                LABEL_START_COL,
                DATA_START_COL + len(YRS) - 1,
                BORDER_TOTAL if k == "root" else BORDER_SUBTOTAL,
            )
    # KCar-style operating sheets. The generic Model sheet remains the canonical
    # engine; these sheets expose the model in the workbook structure expected by
    # the Excel modeling guide and the finished KCar reference workbook.
    module_formula_count = 0
    module_link_count = 0
    module_sheets: dict[str, list[str]] = {}
    if is_kcar:
        module_sheets = _kcar_sheet_nodes(D, display_order)
        for sheet_name, node_ids in module_sheets.items():
            if not node_ids:
                continue
            f_count, l_count = _write_model_sheet(
                wb, sheet_name, sheet_name, node_ids, D, YRS, levels, asts,
                row_map, input_row_map, formula_audit, hist_n,
            )
            module_formula_count += f_count
            module_link_count += l_count

    # Control
    ws_control = wb.create_sheet("Control")
    _setup_sheet(ws_control, "Control")
    _write_year_headers(ws_control, YRS, hist_n)
    _set_model_columns(ws_control, YRS)
    control_nodes = _important_outputs(D, display_order)
    # 열 구성은 _write_year_headers가 이미 잡았다 (F 계정 · G ID · H 단위 · I~ 연도).
    # 예전엔 여기서 B/C에 'Output'/'ID' 헤더를 또 써서 헤더가 두 벌로 겹쳤고,
    # 폭 2짜리 B·C에 계정명을 적어 글자가 잘렸다.
    ws_control.cell(row=HEADER_ROW, column=ID_COL, value="ID").font = FONT_HEADER
    ws_control.cell(row=HEADER_ROW, column=ID_COL).fill = FILL_HEADER
    ws_control.cell(row=HEADER_ROW, column=UNIT_COL, value="Unit").font = FONT_HEADER
    ws_control.cell(row=HEADER_ROW, column=UNIT_COL).fill = FILL_HEADER
    for r, k in enumerate(control_nodes, DATA_START_ROW):
        d = D[k]
        _write_label(ws_control, r, 0, d.get("label", k), bold=(k == "root"))
        ws_control.cell(row=r, column=ID_COL, value=k).font = FONT_ID
        ws_control.cell(row=r, column=UNIT_COL, value=d.get("u", "")).font = FONT_LABEL
        for i in range(len(YRS)):
            cell = ws_control.cell(row=r, column=DATA_START_COL + i, value=f"={_quote_sheet('Model')}!{get_column_letter(DATA_START_COL+i)}{row_map[k]}")
            cell.font = FONT_LINK
            cell.number_format = _guess_number_format(d.get("u", ""), d.get("pct"))
            cell.alignment = ALIGN_RIGHT
        _apply_horizontal_row_border(ws_control, r, 2, DATA_START_COL + len(YRS) - 1)
    # Formula Audit
    ws_audit = wb.create_sheet("Formula Audit")
    _setup_sheet(ws_audit, "Formula Audit")
    for r, row_values in enumerate(formula_audit, 3):
        for cidx, value in enumerate(row_values, 2):
            cell = ws_audit.cell(row=r, column=cidx, value=value)
            if r == 3:
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER
                cell.border = BORDER_HEADER
            else:
                cell.font = FONT_ID if cidx in (3, 6) else FONT_LABEL
        if r > 3:
            _apply_horizontal_row_border(ws_audit, r, 2, 8)
    for col, width in zip(range(2, 9), [14, 24, 30, 10, 10, 60, 16]):
        ws_audit.column_dimensions[get_column_letter(col)].width = width
    # Structure
    ws_structure = wb.create_sheet("Structure")
    _setup_sheet(ws_structure, "Structure")
    headers = ["ID", "Label", "Parent", "Level", "Type", "Formula", "Unit", "Description"]
    for cidx, h in enumerate(headers, 2):
        c = ws_structure.cell(row=3, column=cidx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
    for r, k in enumerate(display_order, 4):
        d = D[k]
        values = [k, d.get("label", ""), d.get("parent", "") or "", levels[k], d.get("type", ""), d.get("formula", "") or "", d.get("u", ""), d.get("desc", "")]
        for cidx, v in enumerate(values, 2):
            ws_structure.cell(row=r, column=cidx, value=v).font = FONT_ID if cidx in (2, 4, 7) else FONT_LABEL
        _apply_horizontal_row_border(ws_structure, r, 2, 9)
    for col, width in zip(range(2, 10), [24, 30, 24, 8, 12, 60, 10, 70]):
        ws_structure.column_dimensions[get_column_letter(col)].width = width
    # Checks
    ws_checks = wb.create_sheet("Checks")
    _setup_sheet(ws_checks, "Checks")
    check_rows = [
        ("Formula cells generated", formula_count + module_formula_count, formula_count > 0),
        ("Input links generated", input_link_count + module_link_count, input_link_count == len(input_order) * len(YRS)),
        ("Operating sheets generated", ", ".join(module_sheets.keys()) if module_sheets else "Generic model only", True),
        ("Computed nodes", len([k for k in D if D[k].get("type") == "computed"]), True),
        ("Input nodes", len(input_order), True),
        ("All computed formulas parsed", "OK", True),
        ("All inputs centralized in Assumptions", "OK", True),
        ("HTML nodes in Structure", len(display_order), len(display_order) == len(D)),
    ]
    ws_checks.cell(row=3, column=2, value="Check").font = FONT_HEADER
    ws_checks.cell(row=3, column=2).fill = FILL_HEADER
    ws_checks.cell(row=3, column=3, value="Result").font = FONT_HEADER
    ws_checks.cell(row=3, column=3).fill = FILL_HEADER
    ws_checks.cell(row=3, column=4, value="Status").font = FONT_HEADER
    ws_checks.cell(row=3, column=4).fill = FILL_HEADER
    for r, (name, result, ok) in enumerate(check_rows, 4):
        ws_checks.cell(row=r, column=2, value=name).font = FONT_LABEL
        ws_checks.cell(row=r, column=3, value=result).font = FONT_FORMULA
        status = ws_checks.cell(row=r, column=4, value="OK" if ok else "FAIL")
        status.font = Font(name="Arial", size=9, bold=True, color="FFFFFF" if not ok else "008000")
        if not ok:
            status.fill = FILL_ERROR
        _apply_horizontal_row_border(ws_checks, r, 2, 4)
    ws_checks.column_dimensions["B"].width = 42
    ws_checks.column_dimensions["C"].width = 24
    ws_checks.column_dimensions["D"].width = 14

    # Metadata
    ws_meta = wb.create_sheet("Metadata")
    _setup_sheet(ws_meta, "Metadata")
    meta = [
        ("Generated by", "financial-modeling/framework/build_excel.py"),
        ("IR version", data.get("version", "?")),
        ("Nodes", len(D)),
        ("Input nodes", len(input_order)),
        ("Computed nodes", len([k for k in D if D[k].get("type") == "computed"])),
        ("Formula cells", formula_count + module_formula_count),
        ("Input link cells", input_link_count + module_link_count),
        ("Years", f"{YRS[0]}–{YRS[-1]} ({len(YRS)}개)"),
        ("Historical years", hist_n),
        ("Forecast years", len(YRS) - hist_n),
        ("Policy", "Inputs are centralized in Assumptions; operating sheets and Model contain live formulas."),
    ]
    for r, (k, v) in enumerate(meta, 3):
        ws_meta.cell(row=r, column=2, value=k).font = FONT_LABEL_BOLD
        ws_meta.cell(row=r, column=3, value=v).font = FONT_LABEL
        _apply_horizontal_row_border(ws_meta, r, 2, 3)
    ws_meta.column_dimensions["B"].width = 24
    ws_meta.column_dimensions["C"].width = 90

    # Workbook-wide finishing
    for ws in wb.worksheets:
        _setup_sheet(ws)
    preferred_order = [
        "Index", "Control", "DCF", "Sales", "Sales Bridge", "Cost", "CapEx,D&A",
        "Labor", "NWC", "BS", "Assumptions", "Model", "Formula Audit", "Structure",
        "Checks", "Metadata",
    ]
    ordered = [wb[s] for s in preferred_order if s in wb.sheetnames]
    ordered.extend([ws for ws in wb.worksheets if ws.title not in preferred_order])
    wb._sheets = ordered
    wb.save(out_path)


# ============================================================
# CLI
# ============================================================

def main() -> int:
    ap = argparse.ArgumentParser(description="BM/HTML 구조 → Excel (수식 박힌 .xlsx) 빌더")
    ap.add_argument("input", help="HTML에서 내보낸 보조 financial_model_*.json")
    ap.add_argument("-o", "--output", default=None, help="출력 .xlsx (기본: <input>.xlsx)")
    args = ap.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        print(f"파일 없음: {in_path}", file=sys.stderr)
        return 1

    out_path = Path(args.output) if args.output else in_path.with_suffix(".xlsx")

    with in_path.open(encoding="utf-8") as f:
        data = json.load(f)

    if "D" not in data or "YRS" not in data:
        print("잘못된 JSON 형식: 'D'와 'YRS' 키 필요", file=sys.stderr)
        return 1

    print(f"읽음: {in_path} ({len(data['D'])}개 노드, {len(data['YRS'])}년)")
    build_workbook(data, out_path)
    print(f"생성: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
